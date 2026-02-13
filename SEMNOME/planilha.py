from abc import ABC, abstractmethod
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import os

"""
Classe para representar uma planilha (worksheet) dentro de uma pasta de trabalho (workbook) Excel.    
"""
class Planilha(ABC):

    # Construtor
    def __init__(self, path: Path, wbOrig_name: str, shOrig_name: str, wbDest: Workbook = None):
        
        # Validações
        if not isinstance(path, Path):
            raise TypeError(f"path deve ser um Path, não {type(path).__name__}")
        
        if not isinstance(wbOrig_name, str):
            raise TypeError(f"wbOrig_name deve ser str, não {type(path).__name__}")
        
        if not isinstance(shOrig_name, str):
            raise TypeError(f"shOrig_name deve ser str, não {type(path).__name__}")
        
        if wbDest is not None and not isinstance(wbDest, Workbook):
            raise TypeError(f"wbDest deve ser um Workbook, não {type(path).__name__}")
        
        # Atributos de construção
        self._path = path
        self._wbOrig_name = wbOrig_name
        self._wsOrig_name = shOrig_name
        self._wbDest = wbDest

        self._wbOrig_path = None
        self._wbOrig = None
        self._wsOrig = None
   
    # Busca uma pasta de trabalho pelo nome no caminho de trabalho
    def findWorkbook(self, path: Path, wb_name: str) -> Path:
        
        # Cria uma lista de strings
        # Cada string é o nome de um arquivo/diretório existente no caminho inspecionado
        files = os.listdir(str(path)) 

        fileNameLower = wb_name.lower()

        # Generator expression
        # Interrompe o laço automaticamente quando encontrada uma correspondência
        # Não encontrando correspondência, retorna None
        foundFile = next((f for f in files if f.lower() == fileNameLower), None)

        if foundFile is None:
            messagebox.showwarning(
                title="Atenção",
                message=f"Arquivo não encontrado: {path}"
            )
            raise FileNotFoundError(f"Arquivo não encontrado: {path}")
        
        return path / foundFile

    # Busca uma aba na pasta de trabalho de origem
    def findWorksheet(self, wb: Workbook, sh_name: str) -> Worksheet: 
        
        # Cria uma lista lowercase dos nomes das abas
        sheet_names_lower = [s.lower() for s in wb.sheetnames]

        try: 
            index = sheet_names_lower.index(sh_name.lower())
        except ValueError:
            raise ValueError(f"Aba não encontrada: {sh_name}")
        
        return wb[wb.sheetnames[index]]

    # Carrega a planilha
    def openWorkbooks(self) -> None:
        
        # Carrega a pasta de trabalho de origem
        try:         
            self._wbOrig_path = self.findWorkbook(self._path, self._wbOrig_name)
            self._wbOrig = load_workbook(self._wbOrig_path)

        except FileNotFoundError:
            raise # Relança a exceção capturada em findWorkbook
        
        # Seleciona a aba da pasta de trabalho de origem
        try: 
            self._wsOrig = self.findWorksheet(self._wbOrig, self._wsOrig_name)
        except ValueError:
            raise # Relança a exceção capturada em findWorksheet

        # Carrega a pasta de trabalho de destino se não foi passada
        if self._wbDest is None:
            
            # Inicializa o Tkinter criando a janela raiz
            root = tk.Tk()

            # Esconde a janela raiz do Tkinter
            root.withdraw()

            try: 
                # Abre uma janela para selecionar a pasta desejada e guarda o caminho
                wbDest_path = filedialog.askopenfilename(
                    title="Selecione o arquivo do orçamento", 
                    filetypes=[("Arquivos Excel", "*.xlsx *.xls *xlsm"), ("Todos os arquivos", "*.*")]
                )

            finally:
                root.destroy()

            if not wbDest_path:
                raise FileNotFoundError("Nenhuma pasta de trabalho de destino foi selecionada.")

            self._wbDest = load_workbook(wbDest_path)

    # Importa os dados da pasta de trabalho de origem para a de destino
    @abstractmethod
    def import_data(self) -> None:
        pass

    # Apaga os dados atuais da pasta de trabalho de destino
    @abstractmethod
    def delete_data(self) -> None:
        pass