from abc import ABC, abstractmethod
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import os

"""
Classe para representar uma planilha (worksheet) dentro de uma pasta de trabalho (workbook) Excel.
"""
class Planilha(ABC):

    # Construtor
    def __init__(self, path: Path, wb_orig_name: str, ws_orig_name: str, wb_dest: Workbook = None):
        
        # Validações
        if not isinstance(path, Path):
            raise TypeError(f"path deve ser um Path, não {type(path).__name__}")
        
        if not isinstance(wb_orig_name, str):
            raise TypeError(f"wb_orig_name deve ser str, não {type(wb_orig_name).__name__}")
        
        if not isinstance(ws_orig_name, str):
            raise TypeError(f"ws_orig_name deve ser str, não {type(ws_orig_name).__name__}")
        
        if wb_dest is not None and not isinstance(wb_dest, Workbook):
            raise TypeError(f"wb_dest deve ser um Workbook, não {type(wb_dest).__name__}")
        
        # Atributos de construção
        self._path = path
        self._wb_orig_name = wb_orig_name
        self._ws_orig_name = ws_orig_name
        self._wb_dest = wb_dest

        self._wb_orig_path = None
        self._wb_orig = None
        self._ws_orig = None

    # Carrega a planilha
    def open_workbooks(self) -> None:
        
        # Carrega a pasta de trabalho de origem
        try:         
            self._wb_orig_path = self.find_workbook(self._path, self._wb_orig_name)
            self._wb_orig = load_workbook(self._wb_orig_path)

        except FileNotFoundError:
            raise # Relança a exceção capturada em find_workbook
        
        # Seleciona a aba da pasta de trabalho de origem
        try: 
            self._ws_orig = self.find_worksheet(self._wb_orig, self._ws_orig_name)
        except ValueError:
            raise # Relança a exceção capturada em find_worksheet

        # Carrega a pasta de trabalho de destino se não foi passada
        if self._wb_dest is None:
            
            # Inicializa o Tkinter criando a janela raiz
            root = tk.Tk()

            # Esconde a janela raiz do Tkinter
            root.withdraw()

            try: 
                # Abre uma janela para selecionar a pasta desejada e guarda o caminho
                wb_dest_path = filedialog.askopenfilename(
                    title="Selecione o arquivo do orçamento", 
                    filetypes=[("Arquivos Excel", "*.xlsx *.xls *xlsm"), ("Todos os arquivos", "*.*")]
                )

            finally:
                root.destroy()

            if not wb_dest_path:
                raise FileNotFoundError("Nenhuma pasta de trabalho de destino foi selecionada.")

            self._wb_dest = load_workbook(wb_dest_path)

    # Busca uma pasta de trabalho pelo nome no caminho de trabalho
    def find_workbook(self, path: Path, wb_name: str) -> Path:
        
        # Cria uma lista de strings
        # Cada string é o nome de um arquivo/diretório existente no caminho inspecionado
        files = os.listdir(str(path)) 

        file_name_lower = wb_name.lower()

        # Generator expression
        # Interrompe o laço automaticamente quando encontrada uma correspondência
        # Não encontrando correspondência, retorna None
        found_file = next((f for f in files if f.lower() == file_name_lower), None)

        if found_file is None:
            messagebox.showwarning(
                title="Atenção",
                message=f"Arquivo não encontrado: {path}"
            )
            raise FileNotFoundError(f"Arquivo não encontrado: {path}")
        
        return path / found_file

    # Busca uma aba na pasta de trabalho de origem
    def find_worksheet(self, wb: Workbook, ws_name: str) -> Worksheet: 
        
        # Cria uma lista lowercase dos nomes das abas
        ws_names_lower = [s.lower() for s in wb.sheetnames]

        try: 
            index = ws_names_lower.index(ws_name.lower())
        except ValueError:
            raise ValueError(f"Aba não encontrada: {ws_name}")
        
        return wb[wb.sheetnames[index]]

    # Importa os dados da pasta de trabalho de origem para a de destino
    @abstractmethod
    def import_data(self) -> None:
        pass

    # Apaga os dados atuais da pasta de trabalho de destino
    @abstractmethod
    def delete_data(self) -> None:
        pass