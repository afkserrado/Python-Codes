from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell
from copy import copy

from Planilha import Planilha

class Cpus(Planilha):
    
    def __init__(self, 
                 path: Path, 
                 wb_orig_name: str, 
                 ws_orig_name: str, 
                 wb_dest: Workbook,
                 ws_dest: Worksheet
    ) -> None:
        
        super().__init__(path, wb_orig_name, ws_orig_name, wb_dest, ws_dest)

    def import_data(self) -> None:
        
        # open_workbooks() deve ter sido chamado antes de invocar este método
        if self._ws_orig is None:
            raise ValueError("Planilha de origem não carregada.")
        
        ws_orig = self._ws_orig
        ws_dest = self._ws_dest
        last_row = self._last_row
        
        if last_row < 5:
            raise ValueError("Não há dados para importar de 'CPUs'.")
        
        # Verifica se a planilha 'CPUs' da pasta de trabalho de destino está vazia
        if self._ws_dest["A7"].value not in (None, ""):
            self.delete_data()
      
        # Intervalo de origem
        min_col, min_row, max_col, max_row = range_boundaries(f"A5:K{last_row}")
        
        # Ponto inicial de destino
        start_row_dest = 6
        start_col_dest = 1 # Coluna A

        # Ponto final de destino
        last_row_dest = start_row_dest + (max_row - min_row)
        last_col_dest = start_col_dest + (max_col - min_col)

        # Insere novas linhas na planilha de destino
        """
        Por exemplo, se origem for A5:K13
        
        n_linhas = 13 - 5 + 1 = 9 linhas a serem copiadas
        linhas_inseridas = 9 - 1 = 8 (não conta a primeira linha, que é do cabeçalho)

        As novas linhas são inseridas sempre a partir da linha 7
        """
        n_linhas = max_row - min_row + 1
        ws_dest.insert_rows(7, n_linhas - 1)

        # Copia os dados com formatação
        for row_orig in ws_orig.iter_rows(
            min_row = min_row,
            max_row = max_row,
            min_col = min_col,
            max_col = max_col
        ):

            for cell_orig in row_orig:

                # Calcula a célula da planilha de destino
                """
                Por exemplo, se origem for A5:K13

                new_row = 5 - 5 + 6 = 6
                new_col = 1 - 1 + 1 = 1

                Linha 6, coluna 1 -> A6

                Quando row for 6

                new_row = 6 - 5 + 6 = 7
                new_col = 1 - 1 + 1 = 1

                Linha 7, coluna 1 -> A7
                """
                row_dest = cell_orig.row - min_row + start_row_dest
                col_dest = cell_orig.column - min_col + start_col_dest

                cell_dest = ws_dest.cell(row=row_dest, column=col_dest)

                # Copia dos dados
                if not isinstance(cell_orig, MergedCell):
                    cell_dest.value = cell_orig.value

                # Copia os estilos
                if cell_orig.has_style:
                    cell_dest.font = copy(cell_orig.font)
                    cell_dest.border = copy(cell_orig.border)
                    cell_dest.fill = copy(cell_orig.fill)
                    cell_dest.number_format = cell_orig.number_format
                    cell_dest.protection = copy(cell_orig.protection)
                    cell_dest.alignment = copy(cell_orig.alignment)
        
        # Mescla as células na planilha de destino
        for merged_range in ws_orig.merged_cells.ranges:
            min_col_m, min_row_m, max_col_m, max_row_m = range_boundaries(str(merged_range))

            # Ignora mesclagens fora do intervalo copiado
            if (
                min_row_m < min_row or
                max_row_m > max_row or
                min_col_m < min_col or
                max_col_m > max_col
            ):
                continue
            
            min_row_dest = min_row_m - min_row + start_row_dest
            max_row_dest = max_row_m - min_row + start_row_dest

            min_col_dest = min_col_m - min_col + start_col_dest
            max_col_dest = max_col_m - min_col + start_col_dest

            range_dest = (
                ws_dest.cell(row=min_row_dest, column=min_col_dest).coordinate
                + ":"
                + ws_dest.cell(row=max_row_dest, column=max_col_dest).coordinate
            )

            ws_dest.merge_cells(range_dest)

        # Ajusta a fonte e o alinhamento global
        for row_dest in ws_dest.iter_rows(
            min_row=start_row_dest,
            max_row=last_row_dest,
            min_col=start_col_dest,
            max_col=last_col_dest
        ):
            for cell_dest in row_dest:
                
                font = copy(cell_dest.font)
                font.name = "Aptos Narrow"
                font.size = 10
                cell_dest.font = font
                
                alignment = copy(cell_dest.alignment)
                alignment.vertical = "center"
                alignment.wrap_text = True
                cell_dest.alignment = alignment

        # Ajusta a largura da coluna D
        ws_dest.column_dimensions["D"].width = 65

        # Aplica filtro
        ws_dest.auto_filter.ref = f"A5:K{last_row_dest}"

    def delete_data(self) -> bool:    
        ws = self._ws_dest
        last_row = self._get_last_row_in_column(ws, "J")

        if last_row >= 7:
            ws.delete_rows(7, last_row - 6)
            return True

        return False