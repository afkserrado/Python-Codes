from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import os

"""
Classe para representar uma planilha (worksheet) dentro de uma pasta de trabalho (workbook) Excel.
"""
class ExcelLoader:

    # Busca e carrega uma pasta de trabalho
    @staticmethod
    def load_workbook(path: Path, wb_name: str) -> Workbook:
        
        files = os.listdir(str(path)) 
        wb_name_lower = wb_name.lower()

        # Generator expression
        # Interrompe o laço automaticamente quando encontrada uma correspondência
        # Não encontrando correspondência, retorna None
        found_file = next(
            (f for f in files if f.lower() == wb_name_lower),
            None
        )

        if found_file is None:
            raise FileNotFoundError(f"Arquivo não encontrado: {path / wb_name}")
        
        wb_path = path / found_file
        return load_workbook(wb_path)

    # Busca e carrega uma aba de uma pasta de trabalho
    @staticmethod
    def load_worksheet(wb: Workbook, ws_name: str) -> Worksheet: 
        for name in wb.sheetnames:
            if name.lower() == ws_name.lower():
                return wb[name]

        raise ValueError(f"Aba não encontrada: {ws_name}")