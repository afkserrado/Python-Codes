from abc import ABC, abstractmethod
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import sys

from ExcelLoader import ExcelLoader

"""
Classe para representar uma planilha (worksheet) dentro de uma pasta de trabalho (workbook) Excel.
"""
class Planilha(ABC):

    # Construtor
    def __init__(self, 
                 path: Path, 
                 wb_orig_name: str, 
                 ws_orig_name: str, 
                 wb_dest: Workbook,
                 ws_dest: Worksheet
    ) -> None:
        
        if not isinstance(path, Path):
            raise TypeError(f"path deve ser um Path, não {type(path).__name__}")
        
        if not isinstance(wb_orig_name, str):
            raise TypeError(f"wb_orig_name deve ser str, não {type(wb_orig_name).__name__}")
        
        if not isinstance(ws_orig_name, str):
            raise TypeError(f"ws_orig_name deve ser str, não {type(ws_orig_name).__name__}")
        
        if not isinstance(wb_dest, Workbook):
            raise TypeError(f"wb_dest deve ser um Workbook, não {type(wb_dest).__name__}")
        
        if not isinstance(ws_dest, Worksheet):
            raise TypeError(f"ws_dest_name deve ser um Worksheet, não {type(ws_dest).__name__}")
        
        self._path = path
        self._wb_orig_name = wb_orig_name
        self._ws_orig_name = ws_orig_name
        self._wb_dest = wb_dest
        self._ws_dest = ws_dest

        self._wb_orig: Workbook | None = None
        self._ws_orig: Worksheet | None = None

        self._last_row: int = 0

    def open_workbooks(self) -> None:
        self._wb_orig = ExcelLoader.load_workbook(self._path, self._wb_orig_name)
        self._ws_orig = ExcelLoader.load_worksheet(self._wb_orig, self._ws_orig_name)
        self._last_row = self._get_last_row_in_column(self._ws_orig, "J")

    # Importa os dados da pasta de trabalho de origem para a de destino
    @abstractmethod
    def import_data(self) -> None:
        pass

    # Apaga os dados atuais da pasta de trabalho de destino
    @abstractmethod
    def delete_data(self) -> bool:
        pass

     # Métodos utilitários
    def _get_last_row_in_column(self, ws: Worksheet, col: str) -> int:
        if not isinstance(col, str) or len(col) != 1:
            raise ValueError("A coluna deve ser uma letra, ex: 'A'.")
        
        col = col.upper()

        # Identifica a última linha com dados na coluna J
        for row in range(ws.max_row, 0, -1):
            if ws[f"{col}{row}"].value is not None:
                return row
            
        return 0

    def ajust_layout(self, path_name: str, ws_dest_name: str) -> None:

        if not sys.platform.startswith("win"):
            return

        import xlwings as xw

        # Abre o Excel via xlwings
        app = xw.App(visible=False) # Não mostra a janela
        
        try:
            wb = xw.Book(path_name)
            ws = wb.sheets[ws_dest_name]

            # Autoajusta colunas e linhas
            ws.autofit()
            ws.range("A:A").column_width = 10
            ws.range("D:D").column_width = 65

            # Ajustar largura da coluna Descrição para 65

            # Define a área de impressão
            last_row = self._last_row + 10
            last_col = ws.api.Cells(6, ws.api.Columns.Count).End(xw.constants.Direction.xlToLeft).Column
            ws.api.PageSetup.PrintArea = ws.range((4,1), (last_row, last_col)).api.Address

            wb.save()
            wb.close()
        
        finally:
            app.quit()